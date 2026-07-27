#!/usr/bin/env python3
"""Local Streamlit app for evaluating Hugging Face candidate cards.

Run from ``Cards/Generation``:

    python -m streamlit run scripts/12_candidate_card_annotation_app.py

The app reads the three ``*_candidate_cards.jsonl`` files from
``data/cards/candidates/huggingface`` and upserts annotations into a local
Excel workbook.
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_INPUT_DIR = PROJECT_DIR / "data" / "cards" / "candidates" / "huggingface"
DEFAULT_OUTPUT_PATH = PROJECT_DIR / "data" / "validation" / "candidate_card_annotations.xlsx"

INPUT_FILES = (
    "llama31_8b_candidate_cards.jsonl",
    "ministral3_8b_candidate_cards.jsonl",
    "qwen35_9b_candidate_cards.jsonl",
)

MODEL_NAMES = {
    "llama31_8b_candidate_cards.jsonl": "Llama 3.1 8B",
    "ministral3_8b_candidate_cards.jsonl": "Ministral 3 8B",
    "qwen35_9b_candidate_cards.jsonl": "Qwen 3.5 9B",
}

SCORE_FIELDS = ("context_relevance", "answer_relevance", "faithfulness")

SCORE_LABELS = {
    "context_relevance": {
        1: "1 · Non rilevante",
        2: "2 · Parzialmente rilevante",
        3: "3 · Pienamente rilevante",
    },
    "answer_relevance": {
        1: "1 · Non rilevante",
        2: "2 · Parzialmente rilevante",
        3: "3 · Pienamente rilevante",
    },
    "faithfulness": {
        1: "1 · Non supportato",
        2: "2 · Parzialmente supportato",
        3: "3 · Pienamente supportato",
    },
}

RUBRIC = (
    {
        "metrica": "context_relevance",
        "score_1": "Il chunk non è pertinente alla descrizione della label EDOS.",
        "score_2": "Il chunk è solo parzialmente pertinente o contiene molto rumore.",
        "score_3": "Il chunk è chiaramente e direttamente pertinente alla label EDOS.",
    },
    {
        "metrica": "answer_relevance",
        "score_1": "L'argument non risponde all'input o è fuori tema.",
        "score_2": "L'argument risponde solo in parte, oppure è troppo generico.",
        "score_3": "L'argument risponde direttamente e in modo specifico all'input.",
    },
    {
        "metrica": "faithfulness",
        "score_1": "L'argument è in contrasto con il chunk o introduce affermazioni non supportate.",
        "score_2": "L'argument è supportato solo in parte dal chunk.",
        "score_3": "Le affermazioni dell'argument sono pienamente supportate dal chunk.",
    },
)

EXCEL_COLUMNS = (
    "card_id",
    "card_order",
    "model",
    "source_file",
    "context_relevance",
    "answer_relevance",
    "faithfulness",
    "notes",
    "annotator",
    "annotated_at",
    "primary_edos_label",
    "edos_description",
    "secondary_edos_labels",
    "status",
    "source_id",
    "source_title",
    "source_publisher",
    "source_year",
    "source_page",
    "source_section",
    "source_url",
    "chunk",
    "reasoning",
    "argument",
    "edos_alignment",
    "retrieval_keywords",
)


def resolve_local_path(value: str, default: Path) -> Path:
    """Resolve a user-entered path relative to the project directory."""
    cleaned = value.strip()
    if not cleaned:
        return default
    path = Path(cleaned).expanduser()
    return path if path.is_absolute() else PROJECT_DIR / path


def extract_edos_description(card: dict[str, Any]) -> str:
    """Extract the EDOS definition stored in the alignment metadata."""
    alignment = str(card.get("edos_alignment") or "")
    match = re.search(r"EDOS definition:\s*(.+?)\s*$", alignment, flags=re.IGNORECASE | re.DOTALL)
    if match:
        return match.group(1).strip()
    return alignment.strip()


@st.cache_data(show_spinner=False)
def load_cards(input_dir_text: str, file_signature: tuple[tuple[str, int], ...]) -> list[dict[str, Any]]:
    """Load and enrich all configured JSONL files.

    ``file_signature`` invalidates Streamlit's cache whenever an input file is
    modified.
    """
    del file_signature
    input_dir = Path(input_dir_text)
    cards: list[dict[str, Any]] = []
    seen_ids: set[str] = set()

    for file_name in INPUT_FILES:
        path = input_dir / file_name
        with path.open("r", encoding="utf-8") as handle:
            for line_number, raw_line in enumerate(handle, start=1):
                if not raw_line.strip():
                    continue
                try:
                    card = json.loads(raw_line)
                except json.JSONDecodeError as exc:
                    raise ValueError(f"JSON non valido in {path.name}, riga {line_number}: {exc}") from exc

                card_id = str(card.get("card_id") or "").strip()
                if not card_id:
                    raise ValueError(f"card_id mancante in {path.name}, riga {line_number}")
                if card_id in seen_ids:
                    raise ValueError(f"card_id duplicato nei file di input: {card_id}")
                seen_ids.add(card_id)

                enriched = dict(card)
                enriched["_model"] = MODEL_NAMES[file_name]
                enriched["_source_file"] = file_name
                enriched["_card_order"] = len(cards) + 1
                enriched["_edos_description"] = extract_edos_description(card)
                cards.append(enriched)

    return cards


def load_annotations(path: Path) -> dict[str, dict[str, Any]]:
    """Read existing annotations, keyed by card_id."""
    if not path.exists():
        return {}
    frame = pd.read_excel(path, sheet_name="annotations", dtype=object)
    if "card_id" not in frame.columns:
        raise ValueError("Il foglio 'annotations' non contiene la colonna card_id.")

    annotations: dict[str, dict[str, Any]] = {}
    for row in frame.to_dict(orient="records"):
        card_id = clean_cell(row.get("card_id"))
        if card_id:
            annotations[str(card_id)] = {
                key: clean_cell(value)
                for key, value in row.items()
            }
    return annotations


def clean_cell(value: Any) -> Any:
    """Convert spreadsheet nulls to empty strings."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return value


def clean_excel_text(value: Any) -> Any:
    """Remove control characters that are illegal in XLSX XML."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def as_score(value: Any) -> int | None:
    """Normalize an Excel/session value to one of the accepted scores."""
    try:
        score = int(value)
    except (TypeError, ValueError):
        return None
    return score if score in (1, 2, 3) else None


def list_to_text(value: Any) -> str:
    if isinstance(value, list):
        return "; ".join(str(item) for item in value)
    return str(value or "")


def build_annotation_row(
    card: dict[str, Any],
    annotator: str,
    scores: dict[str, int],
    notes: str,
) -> dict[str, Any]:
    source = card.get("source") if isinstance(card.get("source"), dict) else {}
    return {
        "card_id": card["card_id"],
        "card_order": card["_card_order"],
        "model": card["_model"],
        "source_file": card["_source_file"],
        "context_relevance": scores["context_relevance"],
        "answer_relevance": scores["answer_relevance"],
        "faithfulness": scores["faithfulness"],
        "notes": notes.strip(),
        "annotator": annotator.strip(),
        "annotated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "primary_edos_label": card.get("primary_edos_label", ""),
        "edos_description": card["_edos_description"],
        "secondary_edos_labels": list_to_text(card.get("secondary_edos_labels")),
        "status": card.get("status", ""),
        "source_id": source.get("source_id", ""),
        "source_title": source.get("title", ""),
        "source_publisher": source.get("publisher", ""),
        "source_year": source.get("year", ""),
        "source_page": source.get("page", ""),
        "source_section": source.get("section", ""),
        "source_url": source.get("url", ""),
        "chunk": card.get("chunk", ""),
        "reasoning": card.get("reasoning", ""),
        "argument": card.get("argument", ""),
        "edos_alignment": card.get("edos_alignment", ""),
        "retrieval_keywords": list_to_text(card.get("retrieval_keywords")),
    }


def save_annotations(path: Path, annotations: dict[str, dict[str, Any]]) -> None:
    """Atomically write annotations and rubric sheets to an XLSX workbook."""
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = path.with_name(f".{path.stem}.tmp.xlsx")

    rows = []
    for annotation in annotations.values():
        row = {column: clean_excel_text(annotation.get(column, "")) for column in EXCEL_COLUMNS}
        rows.append(row)
    rows.sort(key=lambda row: (int(row.get("card_order") or 10**9), str(row.get("card_id") or "")))

    frame = pd.DataFrame(rows, columns=EXCEL_COLUMNS)
    rubric_frame = pd.DataFrame(RUBRIC)

    try:
        with pd.ExcelWriter(temporary_path, engine="openpyxl") as writer:
            frame.to_excel(writer, sheet_name="annotations", index=False)
            rubric_frame.to_excel(writer, sheet_name="rubric", index=False)

            annotations_sheet = writer.book["annotations"]
            annotations_sheet.freeze_panes = "A2"
            annotations_sheet.auto_filter.ref = annotations_sheet.dimensions
            annotations_sheet.row_dimensions[1].height = 28

            header_fill = PatternFill("solid", fgColor="19324D")
            for cell in annotations_sheet[1]:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(vertical="center")

            widths = {
                "A": 34,
                "C": 18,
                "D": 38,
                "E": 19,
                "F": 18,
                "G": 14,
                "H": 36,
                "I": 18,
                "J": 27,
                "K": 42,
                "L": 55,
                "M": 32,
                "O": 30,
                "P": 48,
                "Q": 35,
                "U": 55,
                "V": 90,
                "W": 90,
                "X": 90,
                "Y": 70,
                "Z": 55,
            }
            for column_letter, width in widths.items():
                annotations_sheet.column_dimensions[column_letter].width = width

            for row in annotations_sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            rubric_sheet = writer.book["rubric"]
            rubric_sheet.freeze_panes = "A2"
            for cell in rubric_sheet[1]:
                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
            rubric_sheet.column_dimensions["A"].width = 22
            rubric_sheet.column_dimensions["B"].width = 65
            rubric_sheet.column_dimensions["C"].width = 65
            rubric_sheet.column_dimensions["D"].width = 65
            for row in rubric_sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        temporary_path.replace(path)
    except Exception:
        if temporary_path.exists():
            temporary_path.unlink()
        raise


def inject_css() -> None:
    st.markdown(
        """
        <style>
        .block-container {
            max-width: 1500px;
            padding-top: 2rem;
            padding-bottom: 4rem;
        }
        .hero {
            border-radius: 20px;
            padding: 1.45rem 1.65rem;
            color: #f8fafc;
            background: linear-gradient(125deg, #142f3f 0%, #1e5964 62%, #d17a4a 145%);
            margin-bottom: 1.25rem;
            box-shadow: 0 14px 36px rgba(15, 42, 55, 0.16);
        }
        .hero h1 {
            margin: 0 0 .25rem 0;
            font-size: 2rem;
            letter-spacing: -.03em;
        }
        .hero p {
            margin: 0;
            color: #dbe8ea;
        }
        .metadata {
            border: 1px solid #d9e2e5;
            border-radius: 14px;
            background: #f7faf9;
            padding: .85rem 1rem;
            margin: .35rem 0 1rem 0;
            color: #263b43;
        }
        .eyebrow {
            color: #b95e33;
            font-size: .78rem;
            font-weight: 800;
            letter-spacing: .11em;
            text-transform: uppercase;
            margin-bottom: .3rem;
        }
        .score-panel {
            border: 1px solid #d9e2e5;
            border-radius: 16px;
            padding: .85rem 1rem .25rem 1rem;
            background: #fbfcfb;
            margin-bottom: .75rem;
        }
        div[data-testid="stExpander"] {
            border-color: #d9e2e5;
            border-radius: 14px;
        }
        div[data-testid="stProgress"] > div > div > div {
            background-color: #cf7145;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_metadata(card: dict[str, Any]) -> None:
    source = card.get("source") if isinstance(card.get("source"), dict) else {}

    st.markdown('<div class="eyebrow">Metadati della card</div>', unsafe_allow_html=True)
    first_row = st.columns(4)
    first_row[0].metric("Modello", card["_model"])
    first_row[1].metric("Card globale", f'{card["_card_order"]} / 321')
    first_row[2].metric("Stato", str(card.get("status") or "—"))
    first_row[3].metric("Pagina fonte", str(source.get("page") or "—"))

    st.markdown(
        f"""
        <div class="metadata">
          <strong>Card ID:</strong> {html_escape(card.get("card_id"))}<br>
          <strong>File:</strong> {html_escape(card.get("_source_file"))}<br>
          <strong>Fonte:</strong> {html_escape(source.get("title") or "—")}<br>
          <strong>Editore / anno:</strong> {html_escape(source.get("publisher") or "—")}
          · {html_escape(source.get("year") or "—")}<br>
          <strong>Source ID:</strong> {html_escape(source.get("source_id") or "—")}<br>
          <strong>Label primaria:</strong> {html_escape(card.get("primary_edos_label") or "—")}<br>
          <strong>Label secondarie:</strong>
          {html_escape(list_to_text(card.get("secondary_edos_labels")) or "—")}<br>
          <strong>Retrieval keywords:</strong>
          {html_escape(list_to_text(card.get("retrieval_keywords")) or "—")}
        </div>
        """,
        unsafe_allow_html=True,
    )

    url = str(source.get("url") or "").strip()
    if url:
        st.link_button("Apri la fonte originale ↗", url)


def html_escape(value: Any) -> str:
    import html

    return html.escape(str(value or ""))


def render_content(card: dict[str, Any]) -> None:
    with st.container(border=True):
        st.markdown('<div class="eyebrow">Input · descrizione della EDOS label</div>', unsafe_allow_html=True)
        st.markdown(f"#### {card.get('primary_edos_label') or 'Label non disponibile'}")
        st.write(card["_edos_description"] or "Descrizione non disponibile.")

    with st.container(border=True):
        st.markdown('<div class="eyebrow">Contesto · chunk recuperato</div>', unsafe_allow_html=True)
        st.write(card.get("chunk") or "Chunk non disponibile.")

    with st.container(border=True):
        st.markdown('<div class="eyebrow">Output · argument del modello</div>', unsafe_allow_html=True)
        st.write(card.get("argument") or "Argument non disponibile.")

    with st.expander("Mostra il ragionamento del modello"):
        st.write(card.get("reasoning") or "Ragionamento non disponibile.")

    with st.expander("Mostra il metadata EDOS alignment"):
        st.write(card.get("edos_alignment") or "EDOS alignment non disponibile.")


def score_widget(field: str, title: str, help_text: str, existing_value: Any, card_id: str) -> int | None:
    key = f"{field}::{card_id}"
    if key not in st.session_state:
        st.session_state[key] = as_score(existing_value)

    with st.container(border=True):
        st.markdown(f"**{title}**")
        st.caption(help_text)
        value = st.radio(
            title,
            options=[1, 2, 3],
            index=None,
            format_func=lambda score: SCORE_LABELS[field][score],
            horizontal=True,
            label_visibility="collapsed",
            key=key,
        )
    return as_score(value)


def main() -> None:
    st.set_page_config(
        page_title="Candidate Card Annotation",
        page_icon="◈",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    inject_css()

    st.markdown(
        """
        <div class="hero">
          <h1>Candidate Card Annotation</h1>
          <p>Valutazione locale di rilevanza e fedeltà per le card EDOS.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    with st.sidebar:
        st.header("Sessione")
        annotator = st.text_input(
            "Annotatore",
            placeholder="Nome o iniziali",
            help="Verrà salvato in ogni riga del file Excel.",
        )
        input_dir_text = st.text_input(
            "Cartella JSONL",
            value=str(DEFAULT_INPUT_DIR),
        )
        output_path_text = st.text_input(
            "File Excel",
            value=str(DEFAULT_OUTPUT_PATH),
        )

    input_dir = resolve_local_path(input_dir_text, DEFAULT_INPUT_DIR)
    output_path = resolve_local_path(output_path_text, DEFAULT_OUTPUT_PATH)
    missing_files = [name for name in INPUT_FILES if not (input_dir / name).is_file()]
    if missing_files:
        st.error(
            "File di input mancanti nella cartella selezionata: "
            + ", ".join(missing_files)
        )
        st.stop()

    signature = tuple(
        (name, (input_dir / name).stat().st_mtime_ns)
        for name in INPUT_FILES
    )
    try:
        cards = load_cards(str(input_dir), signature)
        annotations = load_annotations(output_path)
    except Exception as exc:
        st.error(f"Impossibile caricare i dati: {exc}")
        st.stop()

    labels = sorted({str(card.get("primary_edos_label") or "") for card in cards})
    models = [MODEL_NAMES[name] for name in INPUT_FILES]

    with st.sidebar:
        st.divider()
        st.header("Filtri")
        selected_model = st.selectbox("Modello", ["Tutti"] + models)
        selected_label = st.selectbox("EDOS label", ["Tutte"] + labels)
        selected_state = st.selectbox(
            "Stato annotazione",
            ["Tutte", "Da annotare", "Annotate"],
        )

    filtered_cards = []
    for card in cards:
        is_annotated = card["card_id"] in annotations
        if selected_model != "Tutti" and card["_model"] != selected_model:
            continue
        if selected_label != "Tutte" and card.get("primary_edos_label") != selected_label:
            continue
        if selected_state == "Da annotare" and is_annotated:
            continue
        if selected_state == "Annotate" and not is_annotated:
            continue
        filtered_cards.append(card)

    reviewed_count = sum(card["card_id"] in annotations for card in cards)
    with st.sidebar:
        st.divider()
        st.header("Avanzamento")
        st.write(f"**{reviewed_count} / {len(cards)}** card annotate")
        st.progress(reviewed_count / len(cards) if cards else 0.0)
        st.caption(f"{len(filtered_cards)} card corrispondono ai filtri correnti.")

        if output_path.exists():
            st.download_button(
                "Scarica il file Excel",
                data=output_path.read_bytes(),
                file_name=output_path.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        st.caption(f"Salvataggio locale: {output_path}")

    if not filtered_cards:
        if selected_state == "Da annotare" and reviewed_count == len(cards):
            st.success("Tutte le card sono state annotate.")
        else:
            st.info("Nessuna card corrisponde ai filtri selezionati.")
        st.stop()

    position_key = "candidate_card_position"
    if position_key not in st.session_state:
        st.session_state[position_key] = 0
    st.session_state[position_key] = max(
        0,
        min(int(st.session_state[position_key]), len(filtered_cards) - 1),
    )

    with st.sidebar:
        target_position = st.number_input(
            "Vai alla card filtrata",
            min_value=1,
            max_value=len(filtered_cards),
            value=st.session_state[position_key] + 1,
            step=1,
        )
        if int(target_position) - 1 != st.session_state[position_key]:
            st.session_state[position_key] = int(target_position) - 1
            st.rerun()

    card = filtered_cards[st.session_state[position_key]]
    existing = annotations.get(card["card_id"], {})

    flash_message = st.session_state.pop("annotation_flash", "")
    if flash_message:
        st.success(flash_message)

    status_label = "Annotata" if card["card_id"] in annotations else "Da annotare"
    st.caption(
        f"Card filtrata {st.session_state[position_key] + 1} di {len(filtered_cards)} "
        f"· {status_label}"
    )
    render_metadata(card)

    content_column, annotation_column = st.columns([1.65, 1], gap="large")
    with content_column:
        render_content(card)

    with annotation_column:
        st.subheader("Valutazione")
        st.caption("Seleziona un punteggio da 1 a 3 per ciascuna metrica.")

        scores = {
            "context_relevance": score_widget(
                "context_relevance",
                "Context relevance",
                "Quanto il chunk è rilevante rispetto alla descrizione della label EDOS.",
                existing.get("context_relevance"),
                card["card_id"],
            ),
            "answer_relevance": score_widget(
                "answer_relevance",
                "Answer relevance",
                "Quanto l'argument risponde in modo pertinente all'input.",
                existing.get("answer_relevance"),
                card["card_id"],
            ),
            "faithfulness": score_widget(
                "faithfulness",
                "Faithfulness",
                "Quanto l'argument è sostenuto dal chunk, senza aggiunte non giustificate.",
                existing.get("faithfulness"),
                card["card_id"],
            ),
        }

        notes_key = f"notes::{card['card_id']}"
        if notes_key not in st.session_state:
            st.session_state[notes_key] = str(existing.get("notes") or "")
        notes = st.text_area(
            "Note",
            height=150,
            placeholder="Osservazioni libere sull'annotazione…",
            key=notes_key,
        )

        incomplete = [field for field, score in scores.items() if score is None]
        if incomplete:
            st.warning("Completa tutti e tre i punteggi prima di salvare.")

        save_column, save_next_column = st.columns(2)
        save_clicked = save_column.button(
            "Salva",
            type="primary",
            use_container_width=True,
            disabled=bool(incomplete),
        )
        save_next_clicked = save_next_column.button(
            "Salva e prossima →",
            use_container_width=True,
            disabled=bool(incomplete),
        )

        if save_clicked or save_next_clicked:
            annotation = build_annotation_row(
                card,
                annotator,
                {key: int(value) for key, value in scores.items() if value is not None},
                notes,
            )
            annotations[card["card_id"]] = annotation
            try:
                save_annotations(output_path, annotations)
            except PermissionError:
                st.error(
                    "Impossibile aggiornare il file Excel. Se è aperto in Excel, "
                    "chiudilo e riprova."
                )
            except Exception as exc:
                st.error(f"Errore durante il salvataggio: {exc}")
            else:
                st.session_state["annotation_flash"] = (
                    f"Annotazione salvata in {output_path.name}."
                )
                if save_next_clicked and selected_state != "Da annotare":
                    st.session_state[position_key] = min(
                        st.session_state[position_key] + 1,
                        len(filtered_cards) - 1,
                    )
                st.rerun()

    previous_column, spacer_column, next_column = st.columns([1, 2.3, 1])
    if previous_column.button(
        "← Precedente",
        use_container_width=True,
        disabled=st.session_state[position_key] == 0,
    ):
        st.session_state[position_key] -= 1
        st.rerun()
    if next_column.button(
        "Prossima →",
        use_container_width=True,
        disabled=st.session_state[position_key] >= len(filtered_cards) - 1,
    ):
        st.session_state[position_key] += 1
        st.rerun()


if __name__ == "__main__":
    main()
